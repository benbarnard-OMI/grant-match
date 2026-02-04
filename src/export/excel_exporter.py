"""
Excel export functionality for grant matches.

Creates formatted Excel workbooks with multiple sheets:
- Summary sheet with key metrics
- Matches sheet with all data
- Analytics sheet with charts
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


logger = logging.getLogger(__name__)


@dataclass
class ExportConfig:
    """Configuration for Excel export."""
    include_summary: bool = True
    include_analytics: bool = True
    include_raw_data: bool = False
    color_code_scores: bool = True
    auto_width: bool = True


class ExcelExporter:
    """
    Export grant matches to formatted Excel workbook.
    
    Features:
    - Professional formatting with MPART branding
    - Color-coded priority levels
    - Auto-sized columns
    - Multiple sheets (Summary, Matches, Analytics)
    - Charts and visualizations
    """
    
    # Color scheme
    COLORS = {
        'header': '003366',
        'high_priority': 'FFE6E6',
        'medium_priority': 'FFF4E6',
        'low_priority': 'E8F5E9',
        'policy': 'E3F2FD',
        'data': 'F3E5F5',
        'eval': 'E8F5E9'
    }
    
    def __init__(self, config: Optional[ExportConfig] = None):
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "Excel export requires pandas and openpyxl. "
                "Install with: pip install pandas openpyxl"
            )
        
        self.config = config or ExportConfig()
    
    def export(self, matches: List[Dict], 
               output_path: Optional[str] = None,
               metadata: Optional[Dict] = None) -> str:
        """
        Export matches to Excel file.
        
        Args:
            matches: List of match dictionaries
            output_path: Path for output file (auto-generated if None)
            metadata: Optional metadata about the export
            
        Returns:
            Path to exported file
        """
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"data/mpart_matches_{timestamp}.xlsx"
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Create sheets
        if self.config.include_summary:
            self._create_summary_sheet(wb, matches, metadata)
        
        self._create_matches_sheet(wb, matches)
        
        if self.config.include_analytics:
            self._create_analytics_sheet(wb, matches)
        
        # Save
        wb.save(output_path)
        logger.info(f"Excel exported to {output_path}")
        
        return str(output_path)
    
    def _create_summary_sheet(self, wb: Workbook, 
                             matches: List[Dict],
                             metadata: Optional[Dict]):
        """Create summary overview sheet."""
        ws = wb.active
        ws.title = "Summary"
        
        # Title
        ws['A1'] = "MPART @ UIS Grant Match Summary"
        ws['A1'].font = Font(size=18, bold=True, color=self.COLORS['header'])
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date
        ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
        ws['A3'].font = Font(italic=True)
        
        # Statistics
        row = 5
        ws['A' + str(row)] = "Statistics"
        ws['A' + str(row)].font = Font(size=14, bold=True)
        
        row += 2
        stats = self._calculate_stats(matches)
        
        stats_data = [
            ["Metric", "Value"],
            ["Total Matches", stats['total']],
            ["High Priority (≥80)", stats['high_priority']],
            ["Medium Priority (50-79)", stats['medium_priority']],
            ["Low Priority (<50)", stats['low_priority']],
            ["", ""],
            ["Policy Leads", stats['policy_leads']],
            ["Data Leads", stats['data_leads']],
            ["Rural/Eval Leads", stats['eval_leads']],
        ]
        
        for r in stats_data:
            ws.cell(row=row, column=1, value=r[0])
            ws.cell(row=row, column=2, value=r[1])
            
            if r[0] == "Metric":  # Header row
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color=self.COLORS['header'],
                                           end_color=self.COLORS['header'],
                                           fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
            
            row += 1
        
        # Instructions
        row += 3
        ws['A' + str(row)] = "How to Use This Report"
        ws['A' + str(row)].font = Font(size=14, bold=True)
        
        instructions = [
            "",
            "1. Review 'High Priority' matches first (score ≥80)",
            "2. Check the 'Matches' sheet for detailed information",
            "3. Assign leads based on recommended expertise",
            "4. Track deadlines carefully - urgent items are highlighted",
            "5. Update status as you review each opportunity"
        ]
        
        for instruction in instructions:
            row += 1
            ws['A' + str(row)] = instruction
        
        # Auto-width
        if self.config.auto_width:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
    
    def _create_matches_sheet(self, wb: Workbook, matches: List[Dict]):
        """Create detailed matches sheet."""
        ws = wb.create_sheet("Matches")
        
        # Headers
        headers = [
            "Score", "Priority", "Title", "Agency", "Deadline",
            "Days Until", "Lead", "Rationale", "Action", "ID"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, match in enumerate(matches, 2):
            score = match.get('match_score', 0)
            priority = 'High' if score >= 80 else 'Medium' if score >= 50 else 'Low'
            
            # Calculate days until deadline
            days_until = self._calculate_days_until(match.get('deadline'))
            
            lead = match.get('recommended_lead', '')
            lead_display = {
                'mercenary_policy': 'Policy',
                'mercenary_data': 'Data',
                'mercenary_eval': 'Rural/Eval'
            }.get(lead, 'Unassigned')
            
            row_data = [
                score,
                priority,
                match.get('grant_title', ''),
                match.get('agency', ''),
                match.get('deadline', ''),
                days_until,
                lead_display,
                match.get('rationale', ''),
                match.get('recommended_action', ''),
                match.get('grant_id', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # Color code by score
                if self.config.color_code_scores:
                    if score >= 80:
                        cell.fill = PatternFill(start_color=self.COLORS['high_priority'],
                                               fill_type='solid')
                    elif score >= 50:
                        cell.fill = PatternFill(start_color=self.COLORS['medium_priority'],
                                               fill_type='solid')
                
                # Center align score
                if col == 1:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
                
                # Urgent deadline
                if col == 6 and isinstance(days_until, int) and days_until < 14:
                    cell.font = Font(color='FF0000', bold=True)
        
        # Auto-width columns
        if self.config.auto_width:
            ws.column_dimensions['A'].width = 8   # Score
            ws.column_dimensions['B'].width = 12  # Priority
            ws.column_dimensions['C'].width = 50  # Title
            ws.column_dimensions['D'].width = 25  # Agency
            ws.column_dimensions['E'].width = 15  # Deadline
            ws.column_dimensions['F'].width = 12  # Days Until
            ws.column_dimensions['G'].width = 12  # Lead
            ws.column_dimensions['H'].width = 60  # Rationale
            ws.column_dimensions['I'].width = 30  # Action
            ws.column_dimensions['J'].width = 20  # ID
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_analytics_sheet(self, wb: Workbook, matches: List[Dict]):
        """Create analytics sheet with charts."""
        ws = wb.create_sheet("Analytics")
        
        # Score distribution table
        ws['A1'] = "Score Distribution"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "Range"
        ws['B3'] = "Count"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        score_ranges = {
            '90-100': 0,
            '80-89': 0,
            '70-79': 0,
            '60-69': 0,
            '50-59': 0,
            'Below 50': 0
        }
        
        for match in matches:
            score = match.get('match_score', 0)
            if score >= 90:
                score_ranges['90-100'] += 1
            elif score >= 80:
                score_ranges['80-89'] += 1
            elif score >= 70:
                score_ranges['70-79'] += 1
            elif score >= 60:
                score_ranges['60-69'] += 1
            elif score >= 50:
                score_ranges['50-59'] += 1
            else:
                score_ranges['Below 50'] += 1
        
        row = 4
        for range_name, count in score_ranges.items():
            ws.cell(row=row, column=1, value=range_name)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Lead distribution
        row += 2
        ws.cell(row=row, column=1, value="Lead Distribution")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        ws.cell(row=row, column=1, value="Lead Type")
        ws.cell(row=row, column=2, value="Count")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        lead_counts = {'Policy': 0, 'Data': 0, 'Rural/Eval': 0, 'Unassigned': 0}
        for match in matches:
            lead = match.get('recommended_lead', '')
            if lead == 'mercenary_policy':
                lead_counts['Policy'] += 1
            elif lead == 'mercenary_data':
                lead_counts['Data'] += 1
            elif lead == 'mercenary_eval':
                lead_counts['Rural/Eval'] += 1
            else:
                lead_counts['Unassigned'] += 1
        
        row += 1
        for lead_type, count in lead_counts.items():
            ws.cell(row=row, column=1, value=lead_type)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Agency breakdown (top 10)
        row += 2
        ws.cell(row=row, column=1, value="Top Agencies")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        ws.cell(row=row, column=1, value="Agency")
        ws.cell(row=row, column=2, value="Match Count")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        agency_counts = {}
        for match in matches:
            agency = match.get('agency', 'Unknown')
            agency_counts[agency] = agency_counts.get(agency, 0) + 1
        
        top_agencies = sorted(agency_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        row += 1
        for agency, count in top_agencies:
            ws.cell(row=row, column=1, value=agency)
            ws.cell(row=row, column=2, value=count)
            row += 1
    
    def _calculate_stats(self, matches: List[Dict]) -> Dict:
        """Calculate summary statistics."""
        stats = {
            'total': len(matches),
            'high_priority': 0,
            'medium_priority': 0,
            'low_priority': 0,
            'policy_leads': 0,
            'data_leads': 0,
            'eval_leads': 0
        }
        
        for match in matches:
            score = match.get('match_score', 0)
            if score >= 80:
                stats['high_priority'] += 1
            elif score >= 50:
                stats['medium_priority'] += 1
            else:
                stats['low_priority'] += 1
            
            lead = match.get('recommended_lead', '')
            if lead == 'mercenary_policy':
                stats['policy_leads'] += 1
            elif lead == 'mercenary_data':
                stats['data_leads'] += 1
            elif lead == 'mercenary_eval':
                stats['eval_leads'] += 1
        
        return stats
    
    def _calculate_days_until(self, deadline_str: Optional[str]) -> Any:
        """Calculate days until deadline."""
        if not deadline_str:
            return "N/A"
        
        try:
            deadline = datetime.fromisoformat(deadline_str.replace('Z', '+00:00'))
            days = (deadline - datetime.now()).days
            return days if days >= 0 else "Expired"
        except:
            return "Invalid"


def export_matches_to_excel(matches_file: str = "data/mpart_matches.json",
                            output_file: Optional[str] = None) -> str:
    """
    Convenience function to export matches from JSON file to Excel.
    
    Args:
        matches_file: Path to matches JSON file
        output_file: Output Excel file path
        
    Returns:
        Path to exported Excel file
    """
    # Load matches
    with open(matches_file) as f:
        data = json.load(f)
    
    matches = data.get('matches', [])
    metadata = data.get('metadata', {})
    
    # Export
    exporter = ExcelExporter()
    return exporter.export(matches, output_file, metadata)


if __name__ == '__main__':
    # Test export
    import sys
    
    try:
        output = export_matches_to_excel()
        print(f"✓ Excel exported to: {output}")
    except FileNotFoundError:
        print("❌ No matches file found. Run discovery first.")
        sys.exit(1)
    except ImportError as e:
        print(f"❌ {e}")
        sys.exit(1)
